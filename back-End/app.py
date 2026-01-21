from flask import Flask, jsonify, render_template, url_for, redirect, request, flash, Response
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, text, extract
from flask_login import login_user, LoginManager, login_required, logout_user, current_user
from wtforms import *
from wtforms.validators import *
from rutas import *
from clases import *
from sqlalchemy.orm import joinedload
from functools import wraps
from flask_wtf.csrf import generate_csrf
import random
import secrets
import string
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address


login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder a esta página.'
login_manager.login_message_category = 'info'

# Rate Limiter - Protección contra fuerza bruta
limiter = Limiter(
    key_func=get_remote_address,
    app=app,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Exponer csrf_token() como función global de Jinja
app.jinja_env.globals['csrf_token'] = generate_csrf

@login_manager.user_loader
def load_user(user_id):
    # Soportar tanto Club como Member
    if isinstance(user_id, str) and user_id.startswith('member_'):
        member_id = int(user_id.replace('member_', ''))
        return Member.query.get(member_id)
    return Club.query.get(int(user_id))

@app.route('/')
@app.route('/home.html')
def index():
    usuarios = User.query.all()
    return render_template("/noLog/home.html", usuarios=usuarios)

@app.route('/registerform', methods=['GET', 'POST'])
@login_required
def register():
    form = RegistrationForm()
    if request.method == "POST":
        if form.validate_on_submit():
            cedula = form.cedula.data
            name = form.name.data
            telefono = form.telefono.data
            email = form.email.data
            # Verificar si ya existe en ESTE club
            user = User.query.filter(
                User.club_id == current_user.id,
                or_(User.cedula == cedula, User.telefono == telefono, User.email == email)
            ).first()
            if user:
                flash('Ya existe un miembro con esa cédula, teléfono o email.', 'error')
                return redirect(url_for('miembros'))
            new_user = User(cedula=cedula, name=name, telefono=telefono, email=email, club_id=current_user.id)
            db.session.add(new_user)
            db.session.commit()
            flash('Miembro registrado exitosamente.', 'success')
        return redirect(url_for('miembros'))
    return redirect(url_for('miembros', form=form))

@app.route('/miembros.html')
@login_required
def miembros():
    form = RegistrationForm()
    # Filtrar solo miembros de ESTE club
    usuarios = User.query.filter_by(club_id=current_user.id).all()
    return render_template('/logueado/miembros.html', form=form, users=usuarios)



#Funcion que verifica si la cedula esta repetida.
@app.route('/cedula', methods=['POST'])
def verificar_cedula():
    cedula = request.form.get('cedula')

    user = User.query.filter_by(cedula=cedula).first()
    existe_cedula = user is not None

    response = {'existe': existe_cedula}
    return jsonify(response)

# FUNCION QUE SE ENCARGARA DE ELIMINAR SOCIOS DEL CLUB
@app.route('/delete-socio/<int:cedula>', methods=['POST'])
@login_required
def delete_socio(cedula):
    # IMPORTANTE: Verificar que el socio pertenece a ESTE club (seguridad)
    socio = User.query.filter_by(cedula=cedula, club_id=current_user.id).first()
    
    if socio:
        db.session.delete(socio)
        db.session.commit()
        flash('Miembro eliminado correctamente.', 'success')
    else:
        flash('Miembro no encontrado o no pertenece a este club.', 'error')
    
    return redirect(url_for('miembros'))



@app.route('/edit/<cedula>', methods=['GET', 'POST'])
@login_required
def editar_usuario(cedula):
    form = EditForm()
    usuario = User.query.get_or_404(cedula)

    if form.validate_on_submit():
        usuario.cedula = form.cedula.data
        usuario.name = form.name.data
        usuario.telefono =form.telefono.data
        usuario.email = form.email.data
        db.session.commit()
        return redirect(url_for('miembros'))

    form = EditForm(obj=usuario)
    return render_template('/logueado/edit_usuario.html', form=form, usuario=usuario)


@app.route('/registerplanta', methods=['GET', 'POST'])
@login_required
def registerplanta():
    form = PlantForm()
    if request.method == "POST":
        if form.validate_on_submit():
            # Obtener el último ID de planta de ESTE club
            trazabilidad = Trazabilidad.query.filter_by(club_id=current_user.id).order_by(Trazabilidad.idplanta.desc()).first()
            if trazabilidad:
                new_idplanta = trazabilidad.idplanta + 1
            else:
                new_idplanta = 1
            raza = form.raza.data
            Enraizado = form.Enraizado.data
            Riego = form.Riego.data
            paso1 = form.paso1.data
            paso2 = form.paso2.data
            paso3 = form.paso3.data
            floracion = form.floracion.data
            cosecha = form.cosecha.data
            cantidad = int(form.cantidad.data.replace(',', ''))
            observaciones = form.observaciones.data
            new_planta = Trazabilidad(idplanta=new_idplanta, raza=raza, Enraizado=Enraizado, Riego=Riego,
                                       paso1=paso1, paso2=paso2, paso3=paso3, floracion=floracion,
                                       cosecha=cosecha, cantidad=cantidad, observaciones=observaciones,
                                       club_id=current_user.id)
            db.session.add(new_planta)
            db.session.commit()
        return redirect(url_for('trazabilidad'))
    return render_template('/logueado/trazabilidad.html', form=form)

@app.route('/ctrplanta.html')
@app.route('/plantas')
@login_required
def ctrPlantas():
    form = PlantForm()
    # Filtrar plantas de ESTE club
    planta = Trazabilidad.query.filter_by(club_id=current_user.id).all()
    return render_template('/logueado/ctrplanta.html', form=form, planta=planta)

@app.route('/exportar_plantas', methods=['GET'])
@login_required
def exportar_plantas():
    """Exportar datos de trazabilidad de plantas a Excel con filtro de fechas"""
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    # Query base filtrada por club
    query = Trazabilidad.query.filter_by(club_id=current_user.id)
    
    # Aplicar filtros de fecha si se proporcionan (basado en fecha de cosecha)
    if fecha_inicio:
        try:
            inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Trazabilidad.cosecha >= inicio)
        except ValueError:
            pass
    
    if fecha_fin:
        try:
            fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            query = query.filter(Trazabilidad.cosecha <= fin)
        except ValueError:
            pass
    
    plantas = query.all()
    
    # Crear libro Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Control de Plantas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["ID Planta", "Raza", "Riego", "Enraizado", "1er Trasplante", 
               "2do Trasplante", "3er Trasplante", "Floración", "Cosecha", 
               "Cantidad (g)", "Observaciones"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Datos
    for row, planta in enumerate(plantas, 2):
        data = [
            planta.idplanta,
            planta.raza,
            planta.Riego.strftime('%d/%m/%Y') if planta.Riego else '',
            planta.Enraizado.strftime('%d/%m/%Y') if planta.Enraizado else '',
            planta.paso1.strftime('%d/%m/%Y') if planta.paso1 else '',
            planta.paso2.strftime('%d/%m/%Y') if planta.paso2 else '',
            planta.paso3.strftime('%d/%m/%Y') if planta.paso3 else '',
            planta.floracion.strftime('%d/%m/%Y') if planta.floracion else '',
            planta.cosecha.strftime('%d/%m/%Y') if planta.cosecha else '',
            planta.cantidad,
            planta.observaciones or ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # Ajustar ancho de columnas
    column_widths = [12, 20, 12, 12, 15, 15, 15, 12, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Nombre del archivo con fecha
    fecha_actual = datetime.now().strftime('%Y%m%d')
    filename = f"control_plantas_{current_user.username}_{fecha_actual}.xlsx"
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Usar send_file estándar de Flask
    from flask import send_file
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/editplanta/<idplanta>', methods=['GET', 'POST'])
@login_required
def editar_planta(idplanta):
    planta = Trazabilidad.query.get_or_404(idplanta)
    form = PlantForm(obj=planta)

    if form.validate_on_submit():
        planta.raza = form.raza.data
        planta.Enraizado = form.Enraizado.data
        planta.Riego = form.Riego.data
        planta.paso1 = form.paso1.data
        planta.paso2 = form.paso2.data
        planta.paso3 = form.paso3.data
        planta.floracion = form.floracion.data
        planta.cosecha = form.cosecha.data
        if form.cantidad.data:
            planta.cantidad = str(form.cantidad.data).replace(',', '')
        planta.observaciones = form.observaciones.data
        db.session.commit()
        return redirect(url_for('ctrPlantas'))

    return render_template('/logueado/edit_planta.html', form=form, planta=planta)

@app.route('/deleteplanta/<int:idplanta>', methods=['POST'])
@login_required
def delete_planta(idplanta):
    # Verificar que la planta pertenece a ESTE club
    elplant = Trazabilidad.query.filter_by(idplanta=idplanta, club_id=current_user.id).first()
  
    if elplant:
        db.session.delete(elplant)
        db.session.commit()
        flash('Planta eliminada correctamente.', 'success')
    else:
        flash('Planta no encontrada o no pertenece a este club.', 'error')

    return redirect(url_for('ctrPlantas'))


##########################################################################################################################################
@app.route('/registerventas', methods=['GET', 'POST'])
@login_required
def ventosa():
    form = Ventasform()
    if request.method == "POST":
        cedula = form.cedulaVenta.data
        raza = form.razaVenta.data
        cantidad = form.cantVenta.data
        retiro = form.retiro.data
        # Verificar que el usuario pertenece a este club
        usuario = User.query.filter_by(cedula=cedula, club_id=current_user.id).first()
        if usuario:
            new_venta = Ventas(cedula=cedula, raza=raza, cantidad=cantidad, retiro=retiro, club_id=current_user.id)
            db.session.add(new_venta)
            db.session.commit()
        return redirect(url_for('ventas'))
    return render_template('home.html', form=form)

@app.route('/get-username')
def get_username():
    cedula = request.args.get('cedula')
    usuario = User.query.filter_by(cedula=cedula).first()
    if usuario:
        username = usuario.name
    else:
        username = 'Usuario no encontrado'
    return username

@app.route('/delete-venta/<int:idventas>', methods=['POST'])
@login_required
def delete_venta(idventas):
    """Eliminar una venta por su ID"""
    venta = Ventas.query.filter_by(idventas=idventas, club_id=current_user.id).first()
    if venta:
        db.session.delete(venta)
        db.session.commit()
        flash('Venta eliminada correctamente', 'success')
    else:
        flash('Venta no encontrada', 'error')
    return redirect(url_for('ventas'))




@app.route('/ventas.html')
@app.route('/ventas')
@login_required
def ventas():
    form = Ventasform()
    # Filtrar ventas de ESTE club
    ventas = Ventas.query.filter_by(club_id=current_user.id).all()
    return render_template('/logueado/ventas.html', form=form, ventas=ventas)

@app.route('/delete-venta/<int:idventas>')
@login_required
def eliminar_venta(idventas):
    venta = Ventas.query.filter_by(idventas=idventas).first()
    if venta:
        db.session.delete(venta)
        db.session.commit()
    return redirect(url_for('ventas'))

# #######################################
# #########################################################################################################################################################


@app.route('/tabla_ventas')
def tabla_datos():
    datos = obtener_datos()
    return render_template('ventas.html', datos=datos)
def obtener_datos():
    datos = db.session.query(Ventas.idventas, Ventas.cedula, Ventas.raza, Ventas.cantidad, Ventas.retiro).join(User).join(Trazabilidad).all()
    return datos

# #########################################################################################################################################################

@app.route('/prueba', methods=['GET', 'POST'])
def obDatosU():
    # Obtener parámetros de filtro
    tipo_filtro = request.args.get('tipo', 'anual')  # anual, mes, rango
    mes = request.args.get('mes', None)  # 1-12
    desde = request.args.get('desde', None)  # YYYY-MM-DD
    hasta = request.args.get('hasta', None)  # YYYY-MM-DD
    anio = request.args.get('anio', None)  # YYYY
    
    # Query base
    query = Ventas.query
    
    # Aplicar filtros según el tipo
    if tipo_filtro == 'mes' and mes and anio:
        # Filtrar por mes específico
        mes_int = int(mes)
        anio_int = int(anio)
        from datetime import datetime
        fecha_inicio = datetime(anio_int, mes_int, 1)
        if mes_int == 12:
            fecha_fin = datetime(anio_int + 1, 1, 1)
        else:
            fecha_fin = datetime(anio_int, mes_int + 1, 1)
        query = query.filter(Ventas.retiro >= fecha_inicio, Ventas.retiro < fecha_fin)
    
    elif tipo_filtro == 'rango' and desde and hasta:
        # Filtrar por rango de fechas
        from datetime import datetime, timedelta
        fecha_desde = datetime.strptime(desde, '%Y-%m-%d')
        fecha_hasta = datetime.strptime(hasta, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Ventas.retiro >= fecha_desde, Ventas.retiro < fecha_hasta)
    
    elif tipo_filtro == 'anual' and anio:
        # Filtrar por año específico
        from datetime import datetime
        anio_int = int(anio)
        fecha_inicio = datetime(anio_int, 1, 1)
        fecha_fin = datetime(anio_int + 1, 1, 1)
        query = query.filter(Ventas.retiro >= fecha_inicio, Ventas.retiro < fecha_fin)
    
    # Si no hay filtros, devolver todas las ventas (comportamiento original)
    ventas = query.all()

    # Crear lista de datos
    ventas_data = []
    for venta in ventas:
        venta_data = {
            'raza': venta.raza,
            'cantidad': venta.cantidad,
            'retiro': venta.retiro.isoformat() if venta.retiro else None,
        }
        ventas_data.append(venta_data)

    # Devolver los datos en formato JSON
    return jsonify(ventas_data)

# #############################################################################################################################################
# ##############################################################################################################################################

@app.route('/login.html', methods=['POST', 'GET'])
@app.route('/login', methods=['POST', 'GET'])
@limiter.limit("5 per minute", methods=["POST"])  # Protección contra fuerza bruta
def login():
    form = LoginForm()
    
    if current_user.is_authenticated:
        # Redirigir según tipo de usuario ya logueado
        if isinstance(current_user, Member):
            return redirect(url_for('portal_miembro'))
        elif hasattr(current_user, 'is_superuser') and current_user.is_superuser:
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('miembros'))
    
    if request.method == 'POST':
        credential = request.form.get('credential', '').strip()
        password = request.form.get('password', '')
        
        if not credential or not password:
            flash('Por favor ingresa tus credenciales.', 'error')
            return render_template('/noLog/login.html', form=form)
        
        # Primero buscar en Club (por username)
        club_user = Club.query.filter_by(username=credential).first()
        if club_user and club_user.check_password(password):
            login_user(club_user)
            flash('¡Inicio de sesión exitoso!', 'success')
            # Redirigir según tipo
            if club_user.is_superuser:
                return redirect(url_for('admin_dashboard'))
            next_page = request.args.get('next')
            return redirect(next_page or url_for('miembros'))
        
        # Si no se encontró en Club, buscar en Member (por email)
        member_user = Member.query.filter_by(email=credential).first()
        if member_user and member_user.check_password(password):
            login_user(member_user)
            flash('¡Bienvenido!', 'success')
            return redirect(url_for('portal_miembro'))
        
        flash('Usuario o contraseña incorrectos.', 'error')
    
    return render_template('/noLog/login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('index'))


@app.route('/trazabilidad.html')
@login_required
def trazabilidad():
    form = PlantForm()
    return render_template('/logueado/trazabilidad.html', form=form)

@app.route('/graficos.html')
@login_required
def graficos():
    return render_template('/logueado/graficos.html')

# ================================================================================
# RUTAS DE ADMINISTRACIÓN
# ================================================================================

def superuser_required(f):
    """Decorador que requiere que el usuario sea superusuario"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not current_user.is_superuser:
            flash('No tienes permisos para acceder a esta sección.', 'error')
            return redirect(url_for('miembros'))
        return f(*args, **kwargs)
    return decorated_function

def generar_username(longitud=8):
    """Genera un nombre de usuario aleatorio"""
    caracteres = string.ascii_letters + string.digits
    return ''.join(random.choice(caracteres) for _ in range(longitud))

def generar_password(longitud=12):
    """Genera una contraseña aleatoria criptográficamente segura y compleja"""
    if longitud < 12: longitud = 12
    
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    
    while True:
        password = ''.join(secrets.choice(alphabet) for _ in range(longitud))
        if (any(c.islower() for c in password)
                and any(c.isupper() for c in password)
                and any(c.isdigit() for c in password)
                and any(c in "!@#$%" for c in password)):
            return password

@app.route('/admin')
@superuser_required
def admin_dashboard():
    """Dashboard principal de administración"""
    form = LoginForm()  # Para CSRF token
    
    # Estadísticas
    total_clubs = Club.query.filter_by(is_superuser=False).count()
    
    # Clubs de este mes
    from datetime import datetime
    hoy = datetime.now()
    clubs_este_mes = Club.query.filter(
        Club.is_superuser == False,
        extract('year', Club.fecha_creacion) == hoy.year,
        extract('month', Club.fecha_creacion) == hoy.month
    ).count()
    
    # Clubs de este año
    clubs_este_anio = Club.query.filter(
        Club.is_superuser == False,
        extract('year', Club.fecha_creacion) == hoy.year
    ).count()
    
    # Últimos 5 clubs creados
    ultimos_clubs = Club.query.filter_by(is_superuser=False).order_by(Club.fecha_creacion.desc()).limit(5).all()
    
    # Verificar si hay credenciales recién creadas en la sesión
    nuevas_credenciales = None
    
    return render_template('admin/dashboard.html', 
                           form=form,
                           total_clubs=total_clubs,
                           clubs_este_mes=clubs_este_mes,
                           clubs_este_anio=clubs_este_anio,
                           ultimos_clubs=ultimos_clubs,
                           nuevas_credenciales=nuevas_credenciales)

@app.route('/admin/crear-cuenta', methods=['POST'])
@superuser_required
def admin_crear_cuenta():
    """Crear una nueva cuenta de club"""
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    email = request.form.get('email', '').strip()
    
    # Generar credenciales si no se proporcionan
    if not username:
        username = generar_username()
    if not password:
        password = generar_password()
    
    # Verificar si ya existe
    if Club.query.filter_by(username=username).first():
        flash(f'El usuario "{username}" ya existe.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    if email and Club.query.filter_by(email=email).first():
        flash(f'El email "{email}" ya está registrado.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Crear el club
    nuevo_club = Club(username=username, email=email)
    nuevo_club.set_password(password)
    
    db.session.add(nuevo_club)
    db.session.commit()
    
    flash(f'¡Cuenta creada! Usuario: {username} | Contraseña: {password}', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/cuentas')
@superuser_required
def admin_cuentas():
    """Lista todas las cuentas"""
    form = LoginForm()  # Para CSRF token
    clubs = Club.query.order_by(Club.is_superuser.desc(), Club.id.asc()).all()
    return render_template('admin/cuentas.html', form=form, clubs=clubs)

@app.route('/admin/editar-cuenta/<int:id>', methods=['GET', 'POST'])
@superuser_required
def admin_editar_cuenta(id):
    """Editar una cuenta de club"""
    club = Club.query.get_or_404(id)
    
    if club.is_superuser:
        flash('No puedes editar cuentas de administrador.', 'error')
        return redirect(url_for('admin_cuentas'))
    
    if request.method == 'POST':
        nuevo_username = request.form.get('username', '').strip()
        nuevo_email = request.form.get('email', '').strip()
        
        if nuevo_username and nuevo_username != club.username:
            if Club.query.filter_by(username=nuevo_username).first():
                flash('Ese nombre de usuario ya existe.', 'error')
                return redirect(url_for('admin_cuentas'))
            club.username = nuevo_username
        
        if nuevo_email:
            club.email = nuevo_email
        
        db.session.commit()
        flash('Cuenta actualizada correctamente.', 'success')
        return redirect(url_for('admin_cuentas'))
    
    return render_template('admin/editar_cuenta.html', club=club)

@app.route('/admin/resetear-password/<int:id>', methods=['POST'])
@superuser_required
def admin_resetear_password(id):
    """Resetear la contraseña de una cuenta"""
    club = Club.query.get_or_404(id)
    
    if club.is_superuser:
        flash('No puedes resetear contraseñas de administradores.', 'error')
        return redirect(url_for('admin_cuentas'))
    
    nueva_password = generar_password()
    club.set_password(nueva_password)
    db.session.commit()
    
    flash(f'Contraseña reseteada para {club.username}. Nueva contraseña: {nueva_password}', 'success')
    return redirect(url_for('admin_cuentas'))

@app.route('/admin/eliminar-cuenta/<int:id>', methods=['POST'])
@superuser_required
def admin_eliminar_cuenta(id):
    """Eliminar una cuenta de club"""
    club = Club.query.get_or_404(id)
    
    if club.is_superuser:
        flash('No puedes eliminar cuentas de administrador.', 'error')
        return redirect(url_for('admin_cuentas'))
    
    username = club.username
    db.session.delete(club)
    db.session.commit()
    
    flash(f'Cuenta "{username}" eliminada.', 'success')
    return redirect(url_for('admin_cuentas'))

@app.route('/admin/estadisticas')
@superuser_required
def admin_estadisticas():
    """Página de estadísticas detalladas"""
    form = LoginForm()
    
    from datetime import datetime
    hoy = datetime.now()
    
    # Estadísticas por año
    # Estadísticas por año
    stats_anio_query = db.session.query(
        extract('year', Club.fecha_creacion).label('anio'),
        db.func.count(Club.id).label('total')
    ).filter(Club.is_superuser == False).group_by('anio').all()
    
    # Convertir a dict para serialización JSON
    stats_anio = [{'anio': r.anio, 'total': r.total} for r in stats_anio_query]
    
    # Estadísticas por mes del año actual
    stats_mes_query = db.session.query(
        extract('month', Club.fecha_creacion).label('mes'),
        db.func.count(Club.id).label('total')
    ).filter(
        Club.is_superuser == False,
        extract('year', Club.fecha_creacion) == hoy.year
    ).group_by('mes').all()
    
    # Convertir a dict
    stats_mes = [{'mes': r.mes, 'total': r.total} for r in stats_mes_query]
    
    return render_template('admin/estadisticas.html', 
                           form=form,
                           stats_anio=stats_anio,
                           stats_mes=stats_mes)

# ================================================================================
# RUTAS DE MIEMBROS (USUARIOS FINALES)
# ================================================================================

def member_required(f):
    """Decorador que requiere que el usuario sea un Member"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if not isinstance(current_user, Member):
            flash('Acceso solo para miembros registrados.', 'error')
            return redirect(url_for('login_miembro'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/registro-miembro', methods=['GET', 'POST'])
def registro_miembro():
    """Registro de nuevos miembros"""
    form = MemberRegistrationForm()
    
    if form.validate_on_submit():
        cedula = form.cedula.data
        email = form.email.data
        password = form.password.data
        confirm = form.confirm_password.data
        
        # Verificar que las contraseñas coincidan
        if password != confirm:
            flash('Las contraseñas no coinciden.', 'error')
            return render_template('miembro/registro.html', form=form)
        
        # Verificar que la cédula existe en User (miembro de algún club)
        usuario = User.query.filter_by(cedula=cedula).first()
        if not usuario:
            flash('Esta cédula no está registrada en ningún club. Contacta a tu club para registrarte.', 'error')
            return render_template('miembro/registro.html', form=form)
        
        # Verificar que no exista ya una cuenta Member con esa cédula
        if Member.query.filter_by(cedula=cedula).first():
            flash('Ya existe una cuenta con esta cédula.', 'error')
            return render_template('miembro/registro.html', form=form)
        
        # Verificar que el email no esté en uso
        if Member.query.filter_by(email=email).first():
            flash('Este email ya está registrado.', 'error')
            return render_template('miembro/registro.html', form=form)
        
        # Crear el Member (heredando el club_id del User)
        nuevo_miembro = Member(cedula=cedula, email=email, club_id=usuario.club_id)
        nuevo_miembro.set_password(password)
        db.session.add(nuevo_miembro)
        db.session.commit()
        
        flash('¡Cuenta creada exitosamente! Ya puedes iniciar sesión.', 'success')
        return redirect(url_for('login'))
    
    return render_template('miembro/registro.html', form=form)



@app.route('/portal-miembro')
@member_required
def portal_miembro():
    """Portal principal del miembro"""
    # Obtener datos del usuario vinculado
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    
    # Obtener compras
    compras = Ventas.query.filter_by(cedula=current_user.cedula).all()
    total_compras = len(compras)
    total_gramos = sum(c.cantidad for c in compras)
    
    return render_template('miembro/portal.html', 
                           usuario=usuario,
                           total_compras=total_compras,
                           total_gramos=total_gramos)

@app.route('/portal-miembro/compras')
@member_required
def portal_compras():
    """Historial de compras del miembro"""
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    compras = Ventas.query.filter_by(cedula=current_user.cedula).order_by(Ventas.retiro.desc()).all()
    total_gramos = sum(c.cantidad for c in compras)
    
    return render_template('miembro/compras.html',
                           usuario=usuario,
                           compras=compras,
                           total_gramos=total_gramos)

@app.route('/portal-miembro/contacto', methods=['GET', 'POST'])
@member_required
def portal_contacto():
    """Formulario de contacto al club"""
    form = ContactForm()
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    
    if form.validate_on_submit():
        # Aquí podrías enviar email o guardar en BD
        # Por ahora solo mostramos mensaje de éxito
        flash('¡Mensaje enviado correctamente! El club te contactará pronto.', 'success')
        return redirect(url_for('portal_contacto'))
    
    return render_template('miembro/contacto.html', form=form, usuario=usuario)


def calcular_stock(club_id):
    """Calcula el stock disponible por raza considerando reservas de pedidos para un club específico"""
    from sqlalchemy import func
    
    # Sumar gramos cosechados por raza (plantas con fecha de cosecha) DE ESTE CLUB
    cosechado = db.session.query(
        Trazabilidad.raza,
        func.sum(func.cast(Trazabilidad.cantidad, db.Integer)).label('total_cosechado')
    ).filter(
        Trazabilidad.club_id == club_id,
        Trazabilidad.cosecha.isnot(None)
    ).group_by(Trazabilidad.raza).all()
    
    # Sumar gramos vendidos por raza DE ESTE CLUB
    vendido = db.session.query(
        Ventas.raza,
        func.sum(Ventas.cantidad).label('total_vendido')
    ).filter(Ventas.club_id == club_id).group_by(Ventas.raza).all()
    
    # Sumar gramos reservados (pedidos pendientes o coordinados) DE ESTE CLUB
    reservado = db.session.query(
        Pedido.raza,
        func.sum(Pedido.cantidad_solicitada).label('total_reservado')
    ).filter(
        Pedido.club_id == club_id,
        Pedido.estado.in_(['pendiente', 'coordinado'])
    ).group_by(Pedido.raza).all()
    
    # Crear diccionarios
    vendido_dict = {v.raza: v.total_vendido or 0 for v in vendido}
    reservado_dict = {r.raza: r.total_reservado or 0 for r in reservado}
    
    # Calcular stock
    productos = []
    for c in cosechado:
        total = c.total_cosechado or 0
        vendido_qty = vendido_dict.get(c.raza, 0)
        reservado_qty = reservado_dict.get(c.raza, 0)
        disponible = total - vendido_qty - reservado_qty
        
        productos.append({
            'raza': c.raza,
            'total': total,
            'vendido': vendido_qty,
            'reservado': reservado_qty,
            'stock': max(0, disponible)  # Stock disponible para nuevos pedidos
        })
    
    return productos


@app.route('/api/catalogo')
@login_required
def api_catalogo():
    """API que retorna el catálogo con stock"""
    if isinstance(current_user, Member):
        # Obtener club_id del miembro
        usuario = User.query.filter_by(cedula=current_user.cedula).first()
        club_id = usuario.club_id if usuario else None
    else:
        club_id = current_user.id
    
    if not club_id:
        return jsonify([])
    
    productos = calcular_stock(club_id)
    return jsonify(productos)


@app.route('/portal-miembro/catalogo')
@member_required
def portal_catalogo():
    """Catálogo de productos para miembros"""
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    if not usuario:
        flash('Usuario no encontrado.', 'error')
        return redirect(url_for('portal_miembro'))
    
    productos = calcular_stock(usuario.club_id)
    
    # Ordenar por stock (mayor primero)
    productos.sort(key=lambda x: x['stock'], reverse=True)
    
    return render_template('miembro/catalogo.html', usuario=usuario, productos=productos)


@app.route('/portal-miembro/pedido', methods=['GET', 'POST'])
@app.route('/portal-miembro/pedido/<raza>', methods=['GET', 'POST'])
@member_required
def portal_pedido(raza=None):
    """Formulario para coordinar pedido"""
    form = PedidoForm()
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    
    # Si no se especificó raza, redirigir al catálogo
    if not raza:
        return redirect(url_for('portal_catalogo'))
    
    # Obtener stock de esta raza
    productos = calcular_stock(usuario.club_id)
    producto = next((p for p in productos if p['raza'] == raza), None)
    
    if not producto or producto['stock'] <= 0:
        flash('Esta raza no está disponible.', 'error')
        return redirect(url_for('portal_catalogo'))
    
    stock = producto['stock']
    
    if request.method == 'POST':
        cantidad = request.form.get('cantidad', type=int)
        mensaje = request.form.get('mensaje', '')
        
        if not cantidad or cantidad < 1:
            flash('Indica una cantidad válida.', 'error')
        elif cantidad > stock:
            flash(f'Solo hay {stock}g disponibles.', 'error')
        else:
            # Crear pedido con club_id
            pedido = Pedido(
                cedula=current_user.cedula,
                raza=raza,
                cantidad_solicitada=cantidad,
                mensaje=mensaje,
                club_id=usuario.club_id
            )
            db.session.add(pedido)
            db.session.commit()
            
            flash('¡Solicitud enviada! El club te contactará para coordinar.', 'success')
            return redirect(url_for('portal_catalogo'))
    
    return render_template('miembro/pedido.html', 
                           form=form, 
                           usuario=usuario, 
                           raza=raza, 
                           stock=stock)


@app.route('/portal-miembro/mis-pedidos')
@member_required
def portal_mis_pedidos():
    """Lista de pedidos del miembro"""
    usuario = User.query.filter_by(cedula=current_user.cedula).first()
    pedidos = Pedido.query.filter_by(cedula=current_user.cedula).order_by(Pedido.fecha.desc()).all()
    return render_template('miembro/mis_pedidos.html', usuario=usuario, pedidos=pedidos)


# Vista para el club: ver stock de productos
@app.route('/stock')
@login_required
def ver_stock():
    """Vista del club para ver stock actual de productos"""
    if isinstance(current_user, Member):
        return redirect(url_for('portal_miembro'))
    
    productos = calcular_stock(current_user.id)
    # Ordenar por stock disponible (mayor primero)
    productos.sort(key=lambda x: x['stock'], reverse=True)
    
    # Calcular totales
    total_cosechado = sum(p['total'] for p in productos)
    total_vendido = sum(p['vendido'] for p in productos)
    total_reservado = sum(p['reservado'] for p in productos)
    total_disponible = sum(p['stock'] for p in productos)
    
    return render_template('logueado/stock.html', 
                           productos=productos,
                           total_cosechado=total_cosechado,
                           total_vendido=total_vendido,
                           total_reservado=total_reservado,
                           total_disponible=total_disponible)


# Vista para el club: ver pedidos recibidos
@app.route('/pedidos')
@login_required
def ver_pedidos():
    """Vista del club para ver pedidos de miembros"""
    if isinstance(current_user, Member):
        return redirect(url_for('portal_miembro'))
    
    # Filtrar pedidos de ESTE club
    pedidos = Pedido.query.filter_by(club_id=current_user.id).order_by(Pedido.fecha.desc()).all()
    return render_template('logueado/pedidos.html', pedidos=pedidos)


@app.route('/pedido/<int:id>/estado', methods=['POST'])
@login_required
def cambiar_estado_pedido(id):
    """Cambiar estado de un pedido"""
    if isinstance(current_user, Member):
        return redirect(url_for('portal_miembro'))
    
    from datetime import datetime
    
    # Verificar que el pedido pertenece a este club
    pedido = Pedido.query.filter_by(id=id, club_id=current_user.id).first_or_404()
    nuevo_estado = request.form.get('estado')
    estado_anterior = pedido.estado
    
    if nuevo_estado in ['pendiente', 'coordinado', 'completado', 'cancelado']:
        # Si se completa el pedido, crear la venta automáticamente
        if nuevo_estado == 'completado' and estado_anterior != 'completado':
            nueva_venta = Ventas(
                cedula=pedido.cedula,
                raza=pedido.raza,
                cantidad=pedido.cantidad_solicitada,
                retiro=datetime.now(),
                club_id=current_user.id
            )
            db.session.add(nueva_venta)
            flash(f'Pedido completado. Venta de {pedido.cantidad_solicitada}g de {pedido.raza} registrada.', 'success')
        else:
            flash(f'Pedido actualizado a: {nuevo_estado}', 'success')
        
        pedido.estado = nuevo_estado
        db.session.commit()
    
    return redirect(url_for('ver_pedidos'))


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)